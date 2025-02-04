import os
import sys
from openpyxl import load_workbook, Workbook

def list_and_save_sheets(directory, sheet_name):
    """
    Search for a specific sheet name in all Excel files within the given directory and its subdirectories.
    Save each found sheet as a separate sheet in the output Excel file.
    Also, create a recap sheet listing the paths of collected sheets and a summary sheet with all found files and their sheets.

    Args:
        directory (str): The path of the directory to search.
        sheet_name (str): The name of the sheet to extract and save.
    """
    output_workbook = Workbook()
    output_workbook.remove(output_workbook.active)  # Remove default sheet
    recap_sheet = output_workbook.create_sheet(title="Recap")
    recap_sheet.append(["Sheet Name", "File Path"])
    
    summary_sheet = output_workbook.create_sheet(title="Summary")
    summary_sheet.append(["File Path", "File Name", "Sheet Name"])
    
    found = False
    sheet_count = 1

    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith(('.xlsx', '.xlsm')):
                file_path = os.path.join(root, file)
                try:
                    workbook = load_workbook(filename=file_path, read_only=True)
                    for sheet in workbook.sheetnames:
                        summary_sheet.append([file_path, file, sheet])
                    
                    if sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        new_sheet_name = f"{sheet_name}_{sheet_count}"
                        new_sheet = output_workbook.create_sheet(title=new_sheet_name)
                        for row in sheet.iter_rows(values_only=True):
                            new_sheet.append(row)
                        recap_sheet.append([new_sheet_name, file_path])
                        found = True
                        sheet_count += 1
                except Exception as e:
                    print(f"Error reading {file_path}: {e}")
    
    if found:
        output_filename = os.path.join(directory, f"{sheet_name}_collected.xlsx")
        output_workbook.save(output_filename)
        print(f"Sheets named '{sheet_name}' have been saved in {output_filename}")
    else:
        print(f"No sheets named '{sheet_name}' were found in the specified directory.")

def main():
    if len(sys.argv) < 3:
        print("Usage: python script.py <directory> <sheet_name>")
        sys.exit(1)
    
    directory = sys.argv[1]  # Get directory from command-line argument
    sheet_name = sys.argv[2]  # Get sheet name from command-line argument
    list_and_save_sheets(directory, sheet_name)

if __name__ == "__main__":
    main()
